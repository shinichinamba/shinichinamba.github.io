"""openpyxl reader that preserves what pandas would destroy.

pandas is deliberately not used here.  It coerces a column with one blank cell
to float64 (so ``sort_order`` 0 becomes 0.0 and ``2024`` becomes 2024.0), turns
any date-formatted cell into a Timestamp (destroying ``2024-10`` month
precision irrecoverably), and discards ``number_format`` -- which is the only
signal distinguishing a month-accurate date from a day-accurate one.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from .report import Loc, Report


@dataclass(frozen=True)
class RawCell:
    value: object
    number_format: str
    loc: Loc

    @property
    def blank(self) -> bool:
        return self.value is None or (isinstance(self.value, str)
                                      and not self.value.strip())


def date_precision_from_format(number_format: str) -> str | None:
    """Infer a date's intended precision from the cell's display format.

    Quoted literals and locale prefixes such as ``[$-409]`` are stripped first,
    otherwise a stray 'd' inside e.g. "dddd" or a literal would win.
    Returns None when the format carries no date information at all.
    """
    fmt = number_format or ""
    fmt = re.sub(r"\[[^\]]*\]", "", fmt)      # [$-409], [Red], ...
    fmt = re.sub(r'"[^"]*"', "", fmt)         # quoted literals
    fmt = re.sub(r"\\.", "", fmt)             # escaped characters
    fmt = fmt.lower()
    # 'm' is minutes when it follows h/s, but CV dates never carry a time, so
    # treating any m as month is safe here.
    if "d" in fmt:
        return "day"
    if "m" in fmt:
        return "month"
    if "y" in fmt:
        return "year"
    return None


def read_sheet(path: Path, sheet_name: str, rep: Report
               ) -> tuple[list[str], list[dict[str, RawCell]]]:
    wb = openpyxl.load_workbook(path, data_only=True, keep_links=False)
    if sheet_name not in wb.sheetnames:
        rep.error("E-MISSING-SHEET", Loc(path.name, sheet_name),
                  f"sheet {sheet_name!r} not found",
                  f"workbook has: {', '.join(wb.sheetnames)}")
        return [], []
    ws = wb[sheet_name]

    if ws.merged_cells.ranges:
        rep.error("E-MERGED-CELL", Loc(path.name, sheet_name),
                  f"{len(ws.merged_cells.ranges)} merged cell range(s)",
                  "openpyxl reads blanks for all but the top-left cell; "
                  "unmerge them")

    header = [(c.value.strip() if isinstance(c.value, str) else c.value)
              for c in ws[1]]
    while header and header[-1] in (None, ""):
        header.pop()
    header = [h for h in header if h not in (None, "")]

    rows: list[dict[str, RawCell]] = []
    for r in range(2, ws.max_row + 1):
        cells: dict[str, RawCell] = {}
        blank_row = True
        for i, name in enumerate(header, start=1):
            c = ws.cell(row=r, column=i)
            rc = RawCell(c.value, c.number_format or "General",
                         Loc(path.name, sheet_name, r, get_column_letter(i)))
            if not rc.blank:
                blank_row = False
            cells[name] = rc
        if not blank_row:
            rows.append(cells)
    return header, rows
