#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""One-shot initial migration.

Builds data/cv_master.xlsx, data/profile.yml, data/featured_publications.yml,
data/cv_profiles.yml and migration_report.md from the merged initial dataset
in scripts/migration/records.py.

This script is idempotent but is NOT part of the routine build.  After the
first run the workbook is the authoritative copy: edit it, not this script.
Re-running overwrites the workbook, so it refuses to clobber an existing one
unless --force is given.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import yaml

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "scripts"))

from lib.schema import (SHEETS, SHEET_ORDER, physical_columns, bool_columns,  # noqa: E402
                        date_columns, gate_columns, gate_flags, column_map)
from migration.records import DATA  # noqa: E402
from migration.content import (PROFILE, FEATURED, CV_PROFILES,  # noqa: E402
                               DATA_README, MIGRATION_REPORT)

#: Default display policy for gated values.  Amounts stay off the website and
#: off the public short CV; grant numbers are useful on any CV but not on the
#: site.  Every row may override these individually.
GATE_DEFAULTS = {
    "show_amount_web": False,
    "show_amount_cv_short": False,
    "show_amount_cv_full": True,
    "show_grant_number_web": False,
    "show_grant_number_cv_short": True,
    "show_grant_number_cv_full": True,
}

#: which column each gate flag reveals
GATE_OWNER = {
    "show_amount_web": "amount_jpy",
    "show_amount_cv_short": "amount_jpy",
    "show_amount_cv_full": "amount_jpy",
    "show_grant_number_web": "grant_number",
    "show_grant_number_cv_short": "grant_number",
    "show_grant_number_cv_full": "grant_number",
}

HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
NOTE_FILL = PatternFill("solid", fgColor="FFF3C4")


def build_workbook(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for key in SHEET_ORDER:
        sheet = SHEETS[key]
        cols = physical_columns(sheet)
        bools = set(bool_columns(sheet))
        dates = set(date_columns(sheet))
        ws = wb.create_sheet(title=key)

        ws.append(cols)
        for i, name in enumerate(cols, start=1):
            c = ws.cell(row=1, column=i)
            c.font = Font(bold=True)
            c.fill = HEADER_FILL
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

        for rec in DATA[key]:
            row = []
            for name in cols:
                if name in rec:
                    v = rec[name]
                elif name in GATE_DEFAULTS:
                    # Only switch a gate on when the row actually has the
                    # value it would reveal, so the sheet never claims to
                    # display something that is blank.
                    owner = GATE_OWNER[name]
                    v = GATE_DEFAULTS[name] if rec.get(owner) else False
                elif name in bools:
                    v = False
                elif name == "sort_order":
                    v = 0
                else:
                    v = None
                row.append(v)
            ws.append(row)

        # Column widths and, crucially, Text format on every date column so
        # Excel cannot silently promote "2024-10" to a 2024-10-01 datetime.
        for i, name in enumerate(cols, start=1):
            letter = get_column_letter(i)
            width = 12
            if name.endswith(("_ja", "_en")) or name in ("note", "url"):
                width = 34
            elif name.startswith("show_") or name.startswith("visible_"):
                width = 17
            ws.column_dimensions[letter].width = width
            if name in dates:
                for r in range(1, ws.max_row + 1):
                    ws.cell(row=r, column=i).number_format = "@"

    # A short in-workbook reminder of the editing rules.
    ws = wb.create_sheet(title="_README", index=0)
    for i, line in enumerate(DATA_README.strip().splitlines(), start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 100
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.cell(row=1, column=1).fill = NOTE_FILL

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def dump_yaml(path: Path, data, header: str) -> None:
    body = yaml.safe_dump(data, allow_unicode=True, sort_keys=False,
                          default_flow_style=False, width=10 ** 6)
    path.write_text(f"# {header}\n{body}", encoding="utf-8")


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--force", action="store_true",
                    help="overwrite data/cv_master.xlsx if it already exists")
    ap.add_argument("--i-know", action="store_true",
                    help="override the guards that protect spreadsheet edits")
    ap.add_argument("--yaml-only", action="store_true",
                    help="regenerate the YAML inputs only, never the workbook")
    args = ap.parse_args(argv)

    data_dir = REPO / "data"
    xlsx = data_dir / "cv_master.xlsx"
    if xlsx.exists() and not args.force:
        print(f"refusing to overwrite {xlsx.relative_to(REPO)} (use --force)",
              file=sys.stderr)
        return 2

    # --force regenerates the workbook from scripts/migration/records.py and
    # therefore DESTROYS any edit made in Excel. Refuse outright when the file
    # is open, and require --i-know when it has been touched since migration,
    # because by then the spreadsheet is the source of truth, not this script.
    lock = data_dir / "~$cv_master.xlsx"
    if xlsx.exists() and args.force and lock.exists() and not args.i_know:
        print(f"REFUSING: {xlsx.relative_to(REPO)} is open in Excel "
              f"({lock.name} exists).\n"
              f"Close it first. Regenerating would discard your edits.",
              file=sys.stderr)
        return 2
    if xlsx.exists() and args.force and not args.yaml_only:
        # Row counts are the honest signal: once the workbook holds more than
        # the seed data it is the source of truth and must not be rebuilt,
        # whatever --i-know says.
        import openpyxl
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        grown = []
        for key, seed in DATA.items():
            if key not in wb.sheetnames:
                continue
            ws = wb[key]
            rows = sum(1 for r in ws.iter_rows(min_row=2, max_col=1)
                       if r and r[0].value)
            if rows > len(seed):
                grown.append(f"{key} ({rows} rows vs {len(seed)} seeded)")
        wb.close()
        if grown:
            print("REFUSING: the workbook has records this script does not "
                  "know about:\n  " + "\n  ".join(grown) +
                  "\nRebuilding would delete them. Use --yaml-only to "
                  "regenerate just the YAML inputs.", file=sys.stderr)
            return 2

    if xlsx.exists() and args.force and not (args.i_know or args.yaml_only):
        records = Path(__file__).with_name("migration") / "records.py"
        if xlsx.stat().st_mtime > records.stat().st_mtime + 1:
            print(f"REFUSING: {xlsx.relative_to(REPO)} was modified after "
                  f"scripts/migration/records.py.\n"
                  f"The workbook is now the source of truth; regenerating "
                  f"would discard those edits.\n"
                  f"Pass --i-know only if you really want to start over.",
                  file=sys.stderr)
            return 2

    data_dir.mkdir(parents=True, exist_ok=True)
    if not args.yaml_only:
        build_workbook(xlsx)
    dump_yaml(data_dir / "profile.yml", PROFILE,
              "Authored input. Prose, contact details and identifiers.")
    dump_yaml(data_dir / "featured_publications.yml", FEATURED,
              "Authored input. bibkey references _bibliography/*.bib; the "
              "bibliographic data itself is never duplicated here.")
    dump_yaml(data_dir / "cv_profiles.yml", CV_PROFILES,
              "Authored input. CV profile definitions.")
    (data_dir / "README.md").write_text(DATA_README, encoding="utf-8")
    (REPO / "migration_report.md").write_text(MIGRATION_REPORT, encoding="utf-8")

    if args.yaml_only:
        print(f"left {xlsx.relative_to(REPO)} untouched (--yaml-only)")
    else:
        total = sum(len(v) for v in DATA.values())
        print(f"wrote {xlsx.relative_to(REPO)} "
              f"({len(SHEET_ORDER)} sheets, {total} rows)")
    for name in ("profile.yml", "featured_publications.yml", "cv_profiles.yml",
                 "README.md"):
        print(f"wrote data/{name}")
    print("wrote migration_report.md")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
