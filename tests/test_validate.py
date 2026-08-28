"""The validator must refuse bad input loudly."""
import shutil, subprocess, sys, tempfile, unittest
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "scripts"))

import openpyxl  # noqa: E402
from lib.records import load_master  # noqa: E402
from lib.report import Report  # noqa: E402
from lib.validate import Ctx, validate_all  # noqa: E402
from lib.dates import PartialDate, format_range  # noqa: E402


def codes(rep):
    return {i.code for i in rep.issues}


class BrokenWorkbook:
    """Copy the real workbook and corrupt one cell."""

    def __init__(self, sheet, column, row, value):
        self.sheet, self.column, self.row, self.value = sheet, column, row, value

    def __enter__(self):
        self.tmp = Path(tempfile.mkdtemp()) / "cv_master.xlsx"
        shutil.copy(REPO / "data" / "cv_master.xlsx", self.tmp)
        wb = openpyxl.load_workbook(self.tmp)
        ws = wb[self.sheet]
        header = [c.value for c in ws[1]]
        col = header.index(self.column) + 1
        # NB: ws.cell(..., value=None) is a no-op in openpyxl, so assign.
        ws.cell(row=self.row, column=col).value = self.value
        wb.save(self.tmp)
        return self.tmp

    def __exit__(self, *a):
        shutil.rmtree(self.tmp.parent, ignore_errors=True)


def check(path):
    rep = Report()
    master = load_master(path, rep)
    validate_all(Ctx(master, None, {}, [], None), rep)
    return rep


class TestValidation(unittest.TestCase):
    def test_clean_data_passes(self):
        rep = Report()
        master = load_master(REPO / "data" / "cv_master.xlsx", rep)
        validate_all(Ctx(master, None, {}, [], None), rep)
        self.assertEqual(rep.errors, [], rep.render_text())

    def test_duplicate_id(self):
        with BrokenWorkbook("appointments", "id", 3,
                            "utokyo_assistant_professor") as p:
            self.assertIn("E-DUP-ID", codes(check(p)))

    def test_end_before_start(self):
        with BrokenWorkbook("education", "end_date", 2, "2019-01") as p:
            self.assertIn("E-DATE-ORDER", codes(check(p)))

    def test_ongoing_with_end_date_is_only_a_warning(self):
        # A grant can be running now and still have a scheduled end date, so
        # this combination warns rather than failing the build.
        with BrokenWorkbook("appointments", "end_date", 2, "2030-01") as p:
            c = codes(check(p))
            self.assertIn("W-ONGOING-END", c)
            self.assertNotIn("E-ONGOING-END", c)

    def test_bool_column_rejects_string(self):
        # data/README.md requires literal TRUE/FALSE, so even the text
        # "TRUE" is rejected rather than silently rescued.
        for bad in ("true", "TRUE", "yes", "○"):
            with self.subTest(bad=bad):
                with BrokenWorkbook("appointments", "visible_web", 2, bad) as p:
                    self.assertIn("E-BAD-BOOL", codes(check(p)))

    def test_bool_column_rejects_one(self):
        with BrokenWorkbook("appointments", "visible_web", 2, 1) as p:
            self.assertIn("E-BAD-BOOL", codes(check(p)))

    def test_enum_rejects_unknown_value(self):
        with BrokenWorkbook("appointments", "appointment_type", 2,
                            "honorary") as p:
            self.assertIn("E-BAD-ENUM", codes(check(p)))

    def test_unparseable_date(self):
        with BrokenWorkbook("appointments", "start_date", 2, "Oct 2023") as p:
            self.assertIn("E-BAD-DATE", codes(check(p)))

    def test_one_language_missing_warns(self):
        with BrokenWorkbook("appointments", "institution_en", 2, None) as p:
            self.assertIn("W-MISSING-LANG", codes(check(p)))

    def test_required_bilingual_field_needs_at_least_one_side(self):
        with BrokenWorkbook("appointments", "institution_en", 2, None) as p:
            wb = openpyxl.load_workbook(p)
            ws = wb["appointments"]
            hdr = [c.value for c in ws[1]]
            ws.cell(row=2, column=hdr.index("institution_ja") + 1).value = None
            wb.save(p)
            self.assertIn("E-REQUIRED", codes(check(p)))

    def test_missing_column_is_error(self):
        with BrokenWorkbook("appointments", "appointment_type", 1, "bogus") as p:
            self.assertIn("E-MISSING-COLUMN", codes(check(p)))


class TestCliExitCodes(unittest.TestCase):
    def run_cli(self, *args):
        return subprocess.run(
            [sys.executable, str(REPO / "scripts" / "validate_cv_data.py"), *args],
            capture_output=True, text=True)

    def test_clean_run_exits_zero(self):
        r = self.run_cli("--no-check-legacy")
        self.assertEqual(r.returncode, 0, r.stdout + r.stderr)

    def test_list_rules(self):
        r = self.run_cli("--list-rules")
        self.assertEqual(r.returncode, 0)
        self.assertIn("E-DUP-ID", r.stdout)

    def test_explain_dates(self):
        r = self.run_cli("--explain-dates")
        self.assertEqual(r.returncode, 0)
        self.assertIn("precision", r.stdout)


class TestDateOrderPrecision(unittest.TestCase):
    def test_coarse_end_is_not_a_false_positive(self):
        # 2020 as an end date must not look earlier than a 2020-03 start.
        start, end = PartialDate(2020, 3), PartialDate(2020)
        self.assertFalse(end.upper_key() < start.lower_key())

    def test_range_uses_coarser_precision(self):
        s = format_range(PartialDate(2020, 3, 15), PartialDate(2021), False, "en")
        self.assertEqual(s, "2020 – 2021")


if __name__ == "__main__":
    unittest.main(verbosity=2)
